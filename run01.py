import  requests
import  openpyxl
def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)#加载工作簿，打开一个已经存在的excel文件
    sh = wb[sheetname]                 #获取表单
    max_row = sh.max_row               #自动获取最大行数
    case_list = []                     #创建空列表，存储测试用例数据
    for i in range(2,max_row+1):       #取头不取尾，所以最大行数+1
        dict1= dict(                  #嵌套方法，dict(key=value)
        case_id = sh.cell(row=i,column=1).value,     #获取id
        url = sh.cell(row=i,column=5).value,         #获取url
        data = sh.cell(row=i,column=6).value,        #获取单元格中的内容
        expected = sh.cell(row=i,column=7).value     #获取预期结果
        )
        case_list.append(dict1)            #每循环一次，插入到list末尾
    return case_list
#cases = read_data('test_case_api.xlsx','login')   #调用函数，括号（工作簿，sheet）
#print(cases)



#发送接口测试
def api_fun(url,data):    #定义函数，参考
     #URl_login: = 'http://8.129.91.152:8766/futureloan/member/login'
    #json_login = {"mobile_phone": "18699990999", "pwd": "lenmon777"}
     headers = {'X-Lemonban-Media-Type': 'lemonban.v2', 'Content-Type': 'application/json'}
     result = requests.post(url=url, json=data,headers=headers).json()
     return result #设置返回值


#写入excel测试结果
def wirte_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)    #加载工作簿，打开一个已经存在的excel文件
    sh = wb[sheetname]
    sh.cell(row=row,column=column).value = final_result  #直接对单元格进行赋值/修改，写入结果
    wb.save('test_case_api.xlsx')     #保存文档

#eval（函数）---运行被字符串包括的表达式
#str0= '{"mobile_phone":15500000000,"pwd":"12345678996"}'
#dict0 = eval(str0)
#print(dict0)


#接口测试实战，封装成函数
def execute_fun(filename,sheetname):
    cases = read_data(filename,sheetname)
    #print(cases)
    for case in cases:   #依次去访问cases中的元素，保存到定义的case中
    #print(case)
        case_id=case['case_id']
        url = case['url']
        data = eval(case['data'])
        #pirnt(type(data))
        #print(case_id,url,data)

        #获取期望code，msg
        expect = eval(case['expected'])
        expected_code = expect['code']
        expected_msg = expect['msg']
        print('预期结果为code为:{},msg为{}'.format(expected_code,expected_msg))


        #执行接口测试
        real_result = api_fun(url=url,data=data)
        #print(real_result)


        #获取实际结果code、msg
        real_code = real_result['code']
        real_msg = real_result['msg']
        print('实际结果为code为：{}，mas为{}'.format(real_code,real_msg))

        #断言：预期vs实际结果
        if real_result ==expected_code and real_msg == expected_msg:
            print('这{}条用例测试执行通过！'.format(case_id))
            final_re = 'Passed'
        else:
            print('这{}条用例执行不通过！'.format(case_id))
            final_re = 'Failed'
        print('*'*50)

        #写入最终的测试结果
        wirte_result(filename,sheetname,case_id+1,8,final_re)

execute_fun('D:\\pycharm_workspace\\SCB20_web\\python_web\\test_case_api.xlsx', 'register')
        #获取自动化测试需要用到的数据

